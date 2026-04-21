import logging
import os
from openpyxl import load_workbook

from app.models import ProcessNode

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Synthesised L0 and L1 definitions
# ---------------------------------------------------------------------------

# L0 node definitions: id_suffix -> (name, list of domain names)
L0_DEFINITIONS = [
    ("SIP", "Strategy, Infrastructure & Product", ["Market", "Product", "Sales"]),
    ("OPS", "Operations", ["Customer", "Service", "Resource", "Business Partner"]),
    ("ENT", "Enterprise Management", ["Enterprise"]),
]

# Domain name -> L0 id_suffix mapping (built from L0_DEFINITIONS)
DOMAIN_TO_L0: dict[str, str] = {}
for _l0_suffix, _l0_name, _domains in L0_DEFINITIONS:
    for _domain in _domains:
        DOMAIN_TO_L0[_domain] = _l0_suffix

# Domain name -> L1 node id (key with spaces replaced for BusinessPartner)
def _domain_to_l1_id(domain: str) -> str:
    return "L1-" + domain.replace(" ", "")


def _normalise_domain(raw: str | None) -> str | None:
    """Strip trailing ' Domain' suffix and whitespace."""
    if raw is None:
        return None
    s = raw.strip()
    if s.endswith(" Domain"):
        s = s[: -len(" Domain")]
    return s or None


def parse_excel(path: str) -> list[ProcessNode]:
    """Parse the GB921 Excel file and return a list of L0 root nodes with children nested."""

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"GB921 Excel file not found at: {path}\n"
            "Please ensure the file exists before running the parser."
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["eTOM25,5"]

    # ------------------------------------------------------------------
    # 1. Read header row to find column indices by name
    # ------------------------------------------------------------------
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index: dict[str, int] = {
        str(cell).strip(): idx for idx, cell in enumerate(header_row) if cell is not None
    }

    def col(row: tuple, name: str):
        idx = col_index.get(name)
        if idx is None:
            return None
        v = row[idx]
        if isinstance(v, str):
            return v.strip() or None
        return v

    # ------------------------------------------------------------------
    # 2. Build synthesised L0 and L1 nodes (no children yet)
    # ------------------------------------------------------------------
    l0_nodes: dict[str, ProcessNode] = {}
    for suffix, name, domains in L0_DEFINITIONS:
        node_id = f"L0-{suffix}"
        l0_nodes[node_id] = ProcessNode(
            id=node_id,
            name=name,
            level=0,
        )

    # L1 nodes keyed by domain name (normalised, no " Domain" suffix)
    l1_nodes: dict[str, ProcessNode] = {}
    for suffix, _name, domains in L0_DEFINITIONS:
        l0_id = f"L0-{suffix}"
        for domain in domains:
            l1_id = _domain_to_l1_id(domain)
            l1_nodes[domain] = ProcessNode(
                id=l1_id,
                name=domain,
                level=1,
                domain=domain,
                parent_id=l0_id,
            )

    # ------------------------------------------------------------------
    # 3. Read all L2-L7 rows into a flat dict keyed by process_identifier
    # ------------------------------------------------------------------
    data_nodes: dict[str, ProcessNode] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        proc_id = col(row, "Process identifier")
        if proc_id is None:
            continue
        proc_id = str(proc_id).strip()
        if not proc_id:
            continue

        level_raw = col(row, "Level")
        try:
            level = int(level_raw)
        except (TypeError, ValueError):
            logger.warning("Skipping row with invalid level: %r", level_raw)
            continue

        domain_raw = col(row, "Domain")
        domain = _normalise_domain(domain_raw)

        uid_raw = col(row, "UID")
        uid = str(int(uid_raw)) if uid_raw is not None else None

        name_raw = col(row, "Process")
        name = str(name_raw).strip() if name_raw else proc_id

        orig_id_raw = col(row, "Original Process Identifier")
        orig_id = str(orig_id_raw).strip() if orig_id_raw is not None else None

        node = ProcessNode(
            id=proc_id,
            name=name,
            level=level,
            brief_description=col(row, "Brief description"),
            extended_description=col(row, "Extended Description"),
            domain=domain,
            vertical_group=col(row, "Vertical Group"),
            original_id=orig_id,
            uid=uid,
        )
        data_nodes[proc_id] = node

    # ------------------------------------------------------------------
    # 4. Assign parent_id and build children lists
    # ------------------------------------------------------------------

    # Process data nodes sorted by level so parents are assigned before children
    for proc_id, node in data_nodes.items():
        if node.level == 2:
            # Parent is the L1 node for this domain
            domain = node.domain
            if domain and domain in l1_nodes:
                node.parent_id = l1_nodes[domain].id
            else:
                logger.warning(
                    "L2 node %r has unknown domain %r; cannot assign L1 parent", proc_id, domain
                )
        else:
            # Parent is derived by stripping last segment
            parts = proc_id.rsplit(".", 1)
            if len(parts) == 2:
                parent_proc_id = parts[0]
                if parent_proc_id in data_nodes:
                    node.parent_id = data_nodes[parent_proc_id].id
                else:
                    # Fallback: attach to L1 domain
                    domain = node.domain
                    if domain and domain in l1_nodes:
                        node.parent_id = l1_nodes[domain].id
                        logger.warning(
                            "Parent %r not found for node %r; falling back to L1-%s",
                            parent_proc_id, proc_id, domain,
                        )
                    else:
                        logger.warning(
                            "Cannot find parent for node %r (expected parent %r); domain unknown",
                            proc_id, parent_proc_id,
                        )
            else:
                logger.warning("Node %r has no dot-separable parent", proc_id)

    # ------------------------------------------------------------------
    # 5. Wire children into parent nodes
    # ------------------------------------------------------------------

    # Build lookup: all nodes by id
    all_nodes: dict[str, ProcessNode] = {}
    all_nodes.update(l0_nodes)
    for domain, node in l1_nodes.items():
        all_nodes[node.id] = node
    all_nodes.update(data_nodes)

    # Attach L1 nodes to L0 nodes
    for domain, l1_node in l1_nodes.items():
        l0_suffix = DOMAIN_TO_L0.get(domain)
        if l0_suffix:
            l0_id = f"L0-{l0_suffix}"
            l0_nodes[l0_id].children.append(l1_node)

    # Attach data nodes to their parents (process in level order for determinism)
    sorted_data = sorted(data_nodes.values(), key=lambda n: n.level)
    for node in sorted_data:
        if node.parent_id and node.parent_id in all_nodes:
            parent = all_nodes[node.parent_id]
            parent.children.append(node)
        else:
            logger.warning("Node %r has no resolvable parent %r", node.id, node.parent_id)

    wb.close()

    return list(l0_nodes.values())
